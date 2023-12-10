from openpyxl import load_workbook
from openpyxl.styles import Font
import xlsxwriter
import datetime
from googleapiclient.discovery import build
from google.oauth2 import service_account
import csv


PROPERTY_SPREADSHEET_ID = '15ZdwC4U83OvcObbrOYJ_FL2xLKTEYerWwmS-hHAvzHo'
SCOPES = ['https://www.googleapis.com/auth/spreadsheets']
SERVICE_ACCOUNT_FILE = 'keys.json'


class FillForms:
    def __init__(self, locations, prices, links):

        self.locations = locations
        self.prices = prices
        self.links = links
        self.workbook = None
        self.worksheet = None
        self.properties = None
        self.sheet = None
        self.service = None
        self.creds = None
        self.list = None
        self.blue_font = Font(color='000000FF')     # Set blue fonts (for url links)
        # Get the current date and time
        self.time_stamp = datetime.datetime.now().strftime("%d/""%B/""%Y ""%X")  # Format the date and time as dd/Month/yyyy hh:mm:ss (eg 19/February/2022 16:05:33)

    def fill_excel(self):

        try:        # Open the existing workbook and worksheet if they exist
            self.workbook = load_workbook('Rightmove Houses.xlsx')
            self.worksheet = self.workbook['Properties']
        except FileNotFoundError:        # If they don't exist create them and open
            self.workbook = xlsxwriter.Workbook('Rightmove Houses.xlsx')
            self.worksheet = self.workbook.add_worksheet("Properties")

            # Widen columns
            self.worksheet.set_column('A:A', 25)
            self.worksheet.set_column('B:B', 50)
            self.worksheet.set_column('C:C', 10)
            self.worksheet.set_column('D:D', 80)

            # Set columns' names in bold
            format_bold = self.workbook.add_format({'bold': True})
            self.worksheet.write('A1', 'Timestamp', format_bold)
            self.worksheet.write('B1', 'Locations', format_bold)
            self.worksheet.write('C1', 'Prices', format_bold)
            self.worksheet.write('D1', 'Links', format_bold)

            self.workbook.close()

            self.workbook = load_workbook('Rightmove Houses.xlsx')
            self.worksheet = self.workbook['Properties']

        # Identify the number of occupied rows
        last_row = self.worksheet.max_row

        # Fill in  the timestamps
        timestamp_row = last_row + 1
        for _ in range(len(self.locations)):
            self.worksheet[f'A{timestamp_row}'].value = self.time_stamp
            timestamp_row += 1

        # Fill locations in
        location_row = last_row + 1
        for location in self.locations:
            self.worksheet[f'B{location_row}'].value = location
            location_row += 1

        # # Fill prices in
        price_row = last_row + 1
        for price in self.prices:
            self.worksheet[f'C{price_row}'].value = price
            price_row += 1

        # # Fill links in
        link_row = last_row + 1
        for link in self.links:
            self.worksheet[f'D{link_row}'].hyperlink = link
            self.worksheet[f'D{link_row}'].font = self.blue_font
            link_row += 1

        self.workbook.save('Rightmove Houses.xlsx')

    def fill_google_spreadsheet(self):
        # Prepare the Google Sheet
        self.creds = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
        self.service = build('sheets', 'v4', credentials=self.creds)
        self.sheet = self.service.spreadsheets()

        # Prepare the data
        self.properties = []
        for row in range(len(self.locations)):
            self.list = [self.time_stamp, self.locations[row], self.prices[row], self.links[row]]
            self.properties.append(self.list)

        # Fill in the Google Spreadsheet
        self.sheet.values().append(spreadsheetId=PROPERTY_SPREADSHEET_ID, range="Properties!A2"
                                   , valueInputOption="USER_ENTERED", insertDataOption="INSERT_ROWS"
                                   , body={"values": self.properties}).execute()

    def fill_csv(self):
        with open('Rightmove Houses.csv', 'a', newline='') as file:
            writer = csv.writer(file)
            for line in range(len(self.locations)):
                writer.writerow([self.time_stamp, self.locations[line], self.prices[line], self.links[line]])


